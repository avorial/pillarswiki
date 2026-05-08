from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CORE_SHEET_CANDIDATES = {"print sheet", "character sheet", "sheet1"}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def slug(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"^(copy of\s+)", "", value, flags=re.I)
    value = re.sub(r"[^\w\s.-]+", "", value)
    value = re.sub(r"\s+", "-", value.strip())
    value = re.sub(r"-+", "-", value)
    return value.strip(".-") or "character"


def to_number(value: Any) -> int | float | str | None:
    value = clean(value)
    if value == "":
        return None
    try:
        number = float(value)
    except ValueError:
        return value
    return int(number) if number.is_integer() else number


def normalize_key(label: str) -> str:
    label = clean(label).lower()
    label = label.replace('"', "")
    label = re.sub(r"[^a-z0-9]+", "_", label)
    return label.strip("_")


def row_value(rows: list[list[Any]], label: str, offset: int = 2) -> Any:
    label_key = normalize_key(label)
    for row in rows:
        for index, cell in enumerate(row):
            if normalize_key(cell) == label_key:
                for value in row[index + 1 : index + offset + 3]:
                    if clean(value):
                        return value
    return None


def table_pairs(rows: list[list[Any]], start_label: str, stop_labels: set[str], columns: list[tuple[int, int]]) -> dict[str, Any]:
    start = None
    for index, row in enumerate(rows):
        if any(normalize_key(cell) == normalize_key(start_label) for cell in row):
            start = index + 1
            break
    if start is None:
        return {}

    result: dict[str, Any] = {}
    for row in rows[start:]:
        normalized = {normalize_key(cell) for cell in row if clean(cell)}
        if normalized & {normalize_key(label) for label in stop_labels}:
            break
        for name_col, value_col in columns:
            name = clean(row[name_col]) if name_col < len(row) else ""
            value = row[value_col] if value_col < len(row) else ""
            if name and clean(value):
                result[normalize_key(name)] = to_number(value)
    return result


def list_section(
    rows: list[list[Any]],
    start_label: str,
    name_cols: list[int],
    value_cols: list[int] | None = None,
    max_blank_rows: int = 3,
) -> list[Any]:
    start = None
    for index, row in enumerate(rows):
        if any(normalize_key(cell) == normalize_key(start_label) for cell in row):
            start = index + 1
            break
    if start is None:
        return []

    values: list[Any] = []
    blank_rows = 0
    for row in rows[start:]:
        if any(clean(cell).isupper() and len(clean(cell)) > 3 for cell in row):
            # A coarse section-boundary guard for these highly visual sheets.
            if values:
                break
        found = False
        for i, name_col in enumerate(name_cols):
            name = clean(row[name_col]) if name_col < len(row) else ""
            if not name:
                continue
            found = True
            if value_cols:
                value_col = value_cols[i]
                values.append({"name": name, "value": to_number(row[value_col] if value_col < len(row) else "")})
            else:
                values.append(name)
        if found:
            blank_rows = 0
        else:
            blank_rows += 1
            if blank_rows >= max_blank_rows:
                break
    return values


def rows_to_records(rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    header_index = 0
    for index, row in enumerate(rows[:10]):
        if sum(1 for cell in row if clean(cell)) >= 2:
            header_index = index
            break
    headers = [clean(cell) or f"column_{i + 1}" for i, cell in enumerate(rows[header_index])]
    records: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if not any(clean(cell) for cell in row):
            continue
        record = {
            normalize_key(headers[i] if i < len(headers) else f"column_{i + 1}"): to_number(value)
            for i, value in enumerate(row)
            if clean(value)
        }
        if record:
            records.append(record)
    return records


def parse_workbook(payload: dict[str, Any]) -> dict[str, Any]:
    sheets = payload.get("sheets", {})
    core_name = next((name for name in sheets if normalize_key(name) in {normalize_key(s) for s in CORE_SHEET_CANDIDATES}), None)
    if core_name is None:
        core_name = next(iter(sheets), "")
    core = sheets.get(core_name, [])

    def cell(row: int, col: int) -> str:
        return clean(core[row][col]) if row < len(core) and col < len(core[row]) else ""

    character = {
        "source": payload.get("source", {}),
        "folderPath": payload.get("folderPath", []),
        "coreSheet": core_name,
        "character": {
            "chronicle": cell(0, 1),
            "characterName": cell(1, 1),
            "playerName": cell(1, 3),
            "gmailAccount": cell(0, 4) or cell(0, 5) or cell(0, 6),
            "clanCoordinator": cell(1, 5),
            "archetype": cell(3, 1),
            "title": cell(3, 3),
            "characterStartDate": cell(3, 5),
            "generationLabel": cell(5, 1),
            "clan": cell(5, 3),
        },
        "attributes": table_pairs(core, "ATTRIBUTES", {"STATUS"}, [(1, 2)]),
        "backgrounds": table_pairs(core, "BACKGROUNDS", {"MERITS & FLAWS"}, [(8, 9)]),
        "currentPools": table_pairs(core, "CURRENT POOLS", {"BACKGROUND NOTES"}, [(12, 13)]),
        "abilities": table_pairs(core, "ABILITES", {"COMMON POOLS"}, [(1, 2), (3, 4), (5, 6)]),
        "status": list_section(core, "STATUS", [1], None),
        "meritsAndFlaws": list_section(core, "MERITS & FLAWS", [8], None),
        "disciplines": list_section(core, "DISCIPLINES", [12], [13]),
        "ritualsRitaeOther": list_section(core, "RITUALS, RITAE, & OTHER", [12], None),
        "abilitySpecializations": list_section(core, "ABILITES FOCUS", [8], None),
        "rawSheets": {},
    }

    for name, rows in sheets.items():
        if normalize_key(name) == normalize_key(core_name):
            character["rawSheets"][name] = rows
            continue
        character["rawSheets"][name] = {
            "rows": rows,
            "records": rows_to_records(rows),
        }

    return character


def workbook_payload_from_xlsx(path: Path, folder_path: list[str]) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheets: dict[str, list[list[Any]]] = {}
    for sheet in workbook.worksheets:
        rows: list[list[Any]] = []
        for row in sheet.iter_rows(values_only=True):
            values = [clean(cell) for cell in row]
            while values and values[-1] == "":
                values.pop()
            rows.append(values)
        while rows and not any(clean(cell) for cell in rows[-1]):
            rows.pop()
        sheets[sheet.title] = rows
    return {
        "source": {
            "title": path.stem,
            "fileName": path.name,
            "localPath": str(path),
        },
        "folderPath": folder_path,
        "sheets": sheets,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path, help="JSON payload or XLSX character sheet export")
    parser.add_argument("--output-root", type=Path, default=Path("characters"))
    parser.add_argument("--folder-path", default="", help="Slash-separated output folder path, such as Tremere or Minor Clans/Assamite")
    args = parser.parse_args()

    if args.input.suffix.lower() == ".xlsx":
        payload = workbook_payload_from_xlsx(args.input, [part for part in args.folder_path.split("/") if part])
    else:
        payload = json.loads(args.input.read_text(encoding="utf-8"))
    data = parse_workbook(payload)
    folder_path = [slug(part) for part in data.get("folderPath", [])]
    title = data["source"].get("title") or data["character"].get("characterName") or "character"
    output_dir = args.output_root.joinpath(*folder_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{slug(title)}.json"
    output_path.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(output_path)


if __name__ == "__main__":
    main()
